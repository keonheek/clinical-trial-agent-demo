#!/usr/bin/env python3
"""스크리닝 기록 내보내기 — 병원이 증거로 보관하는 형태의 엑셀 + 조회 가능한 SQLite.

두 개를 만든다. 같은 데이터, 다른 용도다.
  · `임상시험_스크리닝기록_YYMMDD.xlsx` — 사람이 읽고 서명하는 문서 (시트 6장)
  · `screening.db` — 같은 내용의 SQLite. "병원이 데이터로 저장해 증거로 남긴다"는 요구는
    엑셀이 아니라 이쪽이 답한다. 엑셀은 DB에서 재생성되는 뷰로 취급한다.

문서 골격은 ICH-GCP 필수문서 체계를 따른다 (스크리닝 로그, 적격성 체크리스트, 질의응답
로그, 정정 로그, 근거 원문). 국내는 식약처 KGCP가 이를 준용하고 실시기관 SOP가 서식을
구체화하므로 병원마다 열 구성은 다르지만 담기는 정보의 골격은 같다.

**단계 구분이 이 파일의 핵심이다.** traces.json은 두 시점을 들고 있다. 초기 판정
(trials[].criteria[].verdict)과, 확인 질문 응답 이후의 재평가(reeval.verdict_changes /
reeval.final_ranking)다. 초기 판정만 내보내면 "질문에 답하기 전 상태"를 최종 판정으로
제시하게 된다 — 10명 중 6명이 재평가에서 판정이 바뀌고, 적격성 등급까지 뒤집히는 환자가
있으므로 임상 독자에게는 단순한 표기 문제가 아니라 오보다. 그래서 모든 시트는 최종 판정을
싣고, 초기 판정은 별도 열과 정정 로그에 남긴다.

LLM을 한 번도 호출하지 않는다. traces.json만 읽고, Vercel 서버리스 경로와 고정된 평가
세트(매칭 정확성)에는 어떤 영향도 주지 않는다.

    python3 export_report.py
    python3 export_report.py --patient S001
    python3 export_report.py --out-dir /some/where --no-db
"""
from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import os
import sqlite3
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from action_policy import apply_trial_level_actions  # noqa: E402
from patient_need import classify_patient_need  # noqa: E402
import ranking  # noqa: E402
from pipeline import decide_eligibility  # noqa: E402
from pipeline import effect_of  # noqa: E402  (판정→효과 매핑의 단일 출처)

ELIG_KO = {"ELIGIBLE": "적격", "UNCERTAIN": "미확정", "INELIGIBLE": "부적격"}
EFFECT_KO = {"PASS": "충족", "FAIL": "탈락", "REVIEW": "확인 필요"}
TYPE_KO = {"inclusion": "포함", "exclusion": "배제"}
ACTION_KO = {"ASK": "환자에게 질문", "TEST": "검사 필요", "CHART": "차트 확인",
             "STOP": "확인 불필요 (이미 부적격)"}
# 자동 재평가는 사람이 아니다. 이 문자열을 "CRC" 같은 사람 직역으로 적으면 서명 개념을
# 위조하는 것이 되므로, 확인자 열은 비워 두고 주체는 이 값으로만 표기한다.
AUTO_ACTOR = "자동 재평가 (AI 파이프라인)"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
REVIEW_FILL = PatternFill("solid", fgColor="FFF6E0")
CHANGED_FILL = PatternFill("solid", fgColor="E8F0FE")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------- 데이터

def _sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _criterion_ids(trials):
    """기준 ID — NCT 뒤 4자리 + 기준 구분 + 구분 내 순번 (9459-EC-02).

    고정된 traces 순서에 대해 결정적이다. 뒤 4자리가 겹치면 ID가 두 시험을 가리키게 되므로,
    겹치는 순간 전체 NCT ID로 떨어진다.

    이 ID가 곧 탈락 사유 코드 역할을 한다. 시험을 가로지르는 의미 기반 코드 체계(EXCL-01 =
    "간기능 수치 초과" 같은)를 새로 만들지 않은 것은 의도적이다 — 그런 분류는 임상적 판단이고,
    이 데이터에는 근거가 없다. 게다가 이 데이터셋은 환자마다 시험 4개가 모두 달라서 시험을
    가로지르는 집계 자체가 성립하지 않는다. 표준 코드 체계는 후속 과제로 남긴다.
    """
    tails = [t.get("nct_id", "")[-4:] for t in trials]
    collide = len(set(tails)) != len(tails)
    out, seen = {}, {}
    for t in trials:
        nct = t.get("nct_id", "")
        stem = nct if collide else nct[-4:]
        for c in t.get("criteria") or []:
            kind = "EC" if c.get("type") == "exclusion" else "IC"
            key = (nct, kind)
            seen[key] = seen.get(key, 0) + 1
            out[(nct, c.get("text", ""))] = f"{stem}-{kind}-{seen[key]:02d}"
    return out


def _apply_reeval(trace):
    """초기 판정 위에 reeval을 얹어 최종 상태를 만든다. traces.json은 건드리지 않는다.

    각 기준에 붙는 것: verdict_initial / verdict_final / effect_final / changed.
    각 시험에 붙는 것: eligibility_final / rank_final / rationale_final.
    """
    reeval = trace.get("reeval") or {}
    changes = {}
    for ch in reeval.get("verdict_changes") or []:
        changes[(ch.get("nct_id"), ch.get("criterion"))] = ch
    final_by_nct = {r.get("nct_id"): r for r in (reeval.get("final_ranking") or [])}

    for t in trace.get("trials", []):
        nct = t.get("nct_id")
        for c in t.get("criteria") or []:
            initial = c.get("verdict")
            c["verdict_initial"] = initial
            ch = changes.get((nct, c.get("text", "")))
            final = ch.get("after") if ch else initial
            c["verdict_final"] = final
            c["effect_final"] = effect_of(c.get("type"), final)
            c["changed"] = bool(ch)
        fr = final_by_nct.get(nct)
        t["eligibility_final"] = (fr or {}).get("eligibility", t.get("eligibility"))
        t["rank_final"] = (fr or {}).get("rank", t.get("rank"))
        t["rationale_final"] = (fr or {}).get("rationale", t.get("rationale"))
        t["n_changed"] = sum(1 for c in t.get("criteria") or [] if c.get("changed"))
    return trace


def load_traces(path=None):
    src = path or os.path.join(HERE, "traces.json")
    with open(src, encoding="utf-8") as f:
        traces = json.load(f)
    return prepare_traces(traces), src


def build_workbook_bytes(traces, src_label, digest):
    """Same workbook as build_workbook, returned as bytes (the /api/export download path)."""
    import io
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_meta(wb, traces, src_label, digest)
    _sheet_screening_log(wb, traces)
    _sheet_checklist(wb, traces)
    _sheet_qa(wb, traces)
    _sheet_corrections(wb, traces)
    _sheet_source(wb, traces)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def prepare_traces(traces):
    """Final-state enrichment shared by the CLI and /api/export (in-memory board state)."""
    for tr in traces:
        apply_trial_level_actions(tr.get("trials", []))
        tr["patient_need"] = classify_patient_need(tr.get("patient_text", ""))
        _apply_reeval(tr)  # verdict_final / effect_final first; the re-rank below reads them
        # Final rank = the order the board shows. traces.json's reeval.final_ranking froze an older
        # rule; api/trace.py re-derives the served order with ranking.rank_trials at request time,
        # so an export that trusted the frozen rank put a different trial at #1 than the screen
        # (demo-video frame check, 2026-08-30). Re-rank a shadow copy built from the FINAL verdicts.
        shadow = []
        for t in tr.get("trials", []):
            crits = [dict(c, verdict=c.get("verdict_final"), effect=c.get("effect_final"))
                     for c in t.get("criteria") or []]
            # eligibility re-derived from the FINAL criteria (pipeline.decide_eligibility), exactly
            # as the answer round does -- reeval.final_ranking's eligibility froze an older rule.
            elig = decide_eligibility(crits)[0] if crits else (t.get("eligibility_final") or "UNCERTAIN")
            t["eligibility_final"] = elig
            shadow.append({"nct_id": t.get("nct_id"), "phase": t.get("phase"), "title": t.get("title"),
                           "eligibility": elig, "trial_intent": t.get("trial_intent"), "criteria": crits})
        ranking.rank_trials(shadow, trust_attached=False, patient_need=tr["patient_need"])
        rank_by = {s_["nct_id"]: s_.get("rank") for s_ in shadow}
        for t in tr.get("trials", []):
            t["rank_final"] = rank_by.get(t.get("nct_id"), t.get("rank_final"))
        tr["_crit_ids"] = _criterion_ids(tr.get("trials", []))
    return traces


def _sorted_trials(trace):
    return sorted(trace.get("trials", []), key=lambda t: (t.get("rank_final") or 99))


def _counts(trial):
    cs = trial.get("criteria") or []
    return (sum(1 for c in cs if c.get("effect_final") == "FAIL"),
            sum(1 for c in cs if c.get("effect_final") == "REVIEW"),
            len(cs))


def _extraction_items(trace):
    """extraction은 dict가 아니라 [{name, value, evidence_quote}] 리스트다."""
    ext = trace.get("extraction")
    if isinstance(ext, list):
        return [(d.get("name", ""), d.get("value", ""), d.get("evidence_quote", ""))
                for d in ext if isinstance(d, dict)]
    if isinstance(ext, dict):
        return [(k, v, "") for k, v in ext.items()]
    return []


def _age_band(trace):
    """생년월일도, 정확한 나이도 싣지 않는다 — 연령대까지만.

    식별 가능성을 줄이는 표준 관행이고, 이 시트가 병원 밖으로 나갈 수 있는 문서이기 때문이다.
    """
    for name, value, _ in _extraction_items(trace):
        if name.strip().lower() != "age":
            continue
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        if digits:
            return f"{int(digits) // 10 * 10}대"
    return ""


def _sex(trace):
    for name, value, _ in _extraction_items(trace):
        if name.strip().lower() == "sex":
            v = str(value).strip().lower()
            return {"male": "M", "female": "F"}.get(v, str(value))
    return ""


def _screening_no(patient_id):
    return f"SCR-{patient_id}"


OUTCOME_KO = {"ELIGIBLE": "PASS", "UNCERTAIN": "PENDING", "INELIGIBLE": "FAIL"}


def _polarity_ko(c):
    """탈락 사유의 극성 — verdict가 아니라 type에서 읽는다.

    verdict=MET은 배제 기준에서는 탈락이고 포함 기준에서는 충족이다. 극성을 verdict로
    판단하면 두 층 구조(verdict / effect)가 막으려던 반전 표기 버그가 그대로 돌아온다.
    """
    if c.get("effect_final") != "FAIL":
        return ""
    return "배제 기준 충족" if c.get("type") == "exclusion" else "포함 기준 미충족"


# --------------------------------------------------------------------------- 엑셀

def _head(ws, headers, widths):
    ws.append(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _finish(ws, ncols):
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _sheet_meta(wb, traces, src, digest):
    ws = wb.create_sheet("메타")
    _head(ws, ["항목", "값"], [26, 116])
    rows = [
        ("생성일", datetime.date.today().isoformat()),
        ("원본 파일", os.path.basename(src)),
        ("원본 SHA256", digest),
        ("환자 수", str(len(traces))),
        ("데이터 성격", "Challenge 제공 합성(synthetic) 환자 시나리오. 실제 환자 기록이 아니며 진료·등록 판단에 사용할 수 없음."),
        ("판정 시점", "모든 시트는 확인 질문 응답 이후의 최종 판정을 싣는다. 초기 판정은 '초기판정' 열과 '정정 로그' 시트에 남는다."),
        ("확인자 열", "비어 있는 것이 정상이다. 자동 파이프라인은 사람이 아니므로 서명란을 자동으로 채우지 않는다. 사람이 검토한 뒤 직접 기입한다."),
        ("문서 골격", "ICH-GCP 필수문서(스크리닝 로그·적격성 체크리스트·질의응답 로그·정정 로그·근거 원문) 구성을 따름. 국내는 식약처 KGCP가 준용."),
        ("CDISC 대응", "탈락사유 시트는 SDTM IE 도메인(Inclusion/Exclusion Criteria Not Met)에, 기준 카탈로그는 TI 도메인에 대응한다. 열 이름은 한국어로 두었다."),
        ("개인정보", "대상자 코드(S001)만 싣는다. 실명·MRN·생년월일은 이 파일에 존재하지 않으며, 실제 운영 시 별도 접근제한 문서로 분리한다."),
        ("보관", "실제 시험에서는 종료 후 최소 3년 보관이 요구된다 (KGCP)."),
    ]
    for k, v in rows:
        ws.append([k, v])
    _finish(ws, 2)


def _sheet_screening_log(wb, traces):
    """스크리닝 로그 — ICH-GCP 8.3.20에 해당하는 문서.

    시험에 접근한 후보를 탈락자까지 포함해 모두 남기는 것이 이 문서의 핵심이다. 등록된
    사람만 남기면 "몇 명을 보고 몇 명이 남았는지"를 사후에 복원할 수 없기 때문이다.
    환자 × 시험 한 행.
    """
    ws = wb.create_sheet("스크리닝 로그")
    _head(ws, ["Screening No.", "환자 ID", "연령대", "성별", "NCT ID", "시험명", "Phase",
               "스크리닝 결과", "최종 판정", "탈락 사유 코드", "탈락 사유", "순위",
               "추천 근거", "판정 변경", "등록 여부"],
          [14, 9, 8, 6, 13, 40, 9, 12, 10, 22, 46, 6, 50, 9, 16])
    for tr in traces:
        ids = tr["_crit_ids"]
        pid = tr.get("patient_id", "")
        band, sex, scr = _age_band(tr), _sex(tr), _screening_no(pid)
        for t in _sorted_trials(tr):
            fails, reviews, n = _counts(t)
            failed = [c for c in (t.get("criteria") or []) if c.get("effect_final") == "FAIL"]
            codes = " · ".join(ids.get((t.get("nct_id"), c.get("text", "")), "") for c in failed)
            texts = " / ".join(c.get("text", "") for c in failed)
            elig = t.get("eligibility_final")
            ws.append([scr, pid, band, sex, t.get("nct_id", ""), t.get("title", ""),
                       t.get("phase") or "NA", OUTCOME_KO.get(elig, "PENDING"),
                       ELIG_KO.get(elig, elig), codes, texts, t.get("rank_final", ""),
                       t.get("rationale_final", ""), t.get("n_changed", 0),
                       "미확정 (데모 범위)"])
            if elig == "INELIGIBLE":
                for c in ws[ws.max_row]:
                    c.fill = FAIL_FILL
    _finish(ws, 17)


def _sheet_checklist(wb, traces):
    """적격성 체크리스트 — 프로토콜 기준 하나하나에 대해 판정·근거·출처를 남기는 문서.

    기준당 1행. 초기판정과 최종판정을 나란히 두는 것이 요점이다 — 확인 질문에 답하기 전과
    후가 다르다는 사실 자체가 이 시스템이 하는 일이고, 그걸 한 열로 뭉개면 재평가가 있었다는
    기록이 사라진다.

    판정 주체는 AI 파이프라인으로 명시하고, 사람이 서명하는 확인자 열은 비워 둔다. 자동
    판정을 사람 서명처럼 적으면 문서가 주장하는 검토 수준을 위조하게 된다.
    """
    ws = wb.create_sheet("적격성 체크리스트")
    _head(ws, ["기준 ID", "환자 ID", "NCT ID", "기준 구분", "기준 원문", "초기판정",
               "최종판정", "결과", "극성", "변경", "근거(EMR 인용)", "판정 근거 설명",
               "다음 행동", "판정 주체", "판정일", "확인자(사람)", "확인일"],
          [13, 9, 13, 9, 52, 11, 11, 11, 15, 7, 34, 46, 18, 22, 11, 13, 11])
    for tr in traces:
        ids = tr["_crit_ids"]
        for t in _sorted_trials(tr):
            for c in t.get("criteria") or []:
                eff = c.get("effect_final")
                ws.append([ids.get((t.get("nct_id"), c.get("text", "")), ""),
                           tr.get("patient_id", ""), t.get("nct_id", ""),
                           TYPE_KO.get(c.get("type"), c.get("type") or ""),
                           c.get("text", ""), c.get("verdict_initial", ""),
                           c.get("verdict_final", ""), EFFECT_KO.get(eff, eff or ""),
                           _polarity_ko(c), "예" if c.get("changed") else "",
                           c.get("evidence", ""), c.get("reasoning", ""),
                           ACTION_KO.get(c.get("action"), c.get("action") or ""),
                           AUTO_ACTOR, (tr.get("generated_at") or "")[:10], "", ""])
                row = ws[ws.max_row]
                if eff == "FAIL":
                    for cell in row:
                        cell.fill = FAIL_FILL
                elif c.get("changed"):
                    for cell in row:
                        cell.fill = CHANGED_FILL
                elif eff == "REVIEW":
                    for cell in row:
                        cell.fill = REVIEW_FILL
    _finish(ws, 17)


def _answer_map(trace):
    """reeval.answers는 {field: answer} 사전이 아니라 [{question, answer}] 리스트다.

    질문 원문으로 맞춘다. 없는 응답을 추측해 채우지 않는다 — 답이 없으면 빈칸이 정답이다.
    """
    out = {}
    ans = (trace.get("reeval") or {}).get("answers")
    if isinstance(ans, list):
        for a in ans:
            if isinstance(a, dict):
                out[(a.get("question") or "").strip()] = a.get("answer")
    elif isinstance(ans, dict):
        out.update({str(k).strip(): v for k, v in ans.items()})
    return out


def _sheet_qa(wb, traces):
    ws = wb.create_sheet("질의응답 이력")
    _head(ws, ["환자 ID", "확인 대상 필드", "확인 질문", "질문 사유", "응답"],
          [10, 22, 60, 52, 40])
    for tr in traces:
        answers = _answer_map(tr)
        for q in tr.get("questions") or []:
            ans = answers.get((q.get("question") or "").strip())
            if isinstance(ans, (dict, list)):
                ans = json.dumps(ans, ensure_ascii=False)
            ws.append([tr.get("patient_id", ""), q.get("field", ""), q.get("question", ""),
                       q.get("why", ""), "" if ans is None else str(ans)])
    _finish(ws, 5)


def _sheet_corrections(wb, traces):
    ws = wb.create_sheet("정정 로그")
    _head(ws, ["일자", "환자 ID", "기준 ID", "NCT ID", "정정 대상 기준", "정정 전",
               "정정 후", "사유", "확인자"],
          [12, 10, 11, 13, 52, 12, 12, 34, 22])
    for tr in traces:
        ids = tr["_crit_ids"]
        when = (tr.get("generated_at") or "")[:10]
        changes = ((tr.get("reeval") or {}).get("verdict_changes")) or []
        if not changes:
            ws.append([when, tr.get("patient_id", ""), "", "", "없음 — 재평가에서 바뀐 판정 없음",
                       "", "", "", ""])
            continue
        for ch in changes:
            ws.append([when, tr.get("patient_id", ""),
                       ids.get((ch.get("nct_id"), ch.get("criterion", "")), ""),
                       ch.get("nct_id", ""), ch.get("criterion", ""),
                       ch.get("before", ""), ch.get("after", ""),
                       "확인 질문 응답 반영", AUTO_ACTOR])
    _finish(ws, 9)


def _sheet_source(wb, traces):
    """근거 원문 — 구조화된 판정이 어느 문장에서 나왔는지 되짚을 수 있게 하는 시트."""
    ws = wb.create_sheet("근거 원문")
    _head(ws, ["환자 ID", "항목", "값", "원문 인용"], [10, 22, 56, 56])
    for tr in traces:
        pid = tr.get("patient_id", "")
        ws.append([pid, "환자 기술(원문)", tr.get("patient_text", ""), ""])
        for name, value, quote in _extraction_items(tr):
            ws.append([pid, name, "" if value is None else str(value), quote])
    _finish(ws, 4)


def build_workbook(traces, src, digest, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_meta(wb, traces, src, digest)
    _sheet_screening_log(wb, traces)
    _sheet_checklist(wb, traces)
    _sheet_qa(wb, traces)
    _sheet_corrections(wb, traces)
    _sheet_source(wb, traces)
    wb.save(out_path)
    return {ws.title: ws.max_row - 1 for ws in wb.worksheets}


# --------------------------------------------------------------------------- SQLite

SCHEMA = """
DROP TABLE IF EXISTS patients;
DROP TABLE IF EXISTS trials;
DROP TABLE IF EXISTS screening_records;
DROP TABLE IF EXISTS criteria_evaluations;
DROP TABLE IF EXISTS qa_log;
DROP TABLE IF EXISTS corrections;
CREATE TABLE patients (
  patient_id TEXT PRIMARY KEY, age_band TEXT, sex TEXT, need TEXT, need_confidence TEXT,
  patient_text TEXT);
CREATE TABLE trials (
  nct_id TEXT PRIMARY KEY, title TEXT, phase TEXT);
CREATE TABLE screening_records (
  screening_no TEXT, patient_id TEXT, age_band TEXT, sex TEXT, nct_id TEXT,
  outcome TEXT, eligibility_final TEXT, fail_codes TEXT, rank_final INTEGER,
  n_fail INTEGER, n_review INTEGER, n_criteria INTEGER, n_changed INTEGER,
  rationale TEXT, enrollment_status TEXT, PRIMARY KEY (patient_id, nct_id));
CREATE TABLE criteria_evaluations (
  criterion_id TEXT, patient_id TEXT, nct_id TEXT, criterion_text TEXT, criterion_type TEXT,
  verdict_initial TEXT, verdict_final TEXT, effect_final TEXT, polarity TEXT, changed INTEGER,
  evidence TEXT, reasoning TEXT, next_action TEXT, decided_by TEXT, decided_at TEXT,
  confirmed_by TEXT, confirmed_at TEXT,
  PRIMARY KEY (patient_id, criterion_id));
CREATE TABLE qa_log (
  patient_id TEXT, field TEXT, question TEXT, why TEXT, answer TEXT);
CREATE TABLE corrections (
  changed_on TEXT, patient_id TEXT, criterion_id TEXT, nct_id TEXT, criterion_text TEXT,
  verdict_before TEXT, verdict_after TEXT, reason TEXT, actor TEXT);
CREATE INDEX idx_crit_effect ON criteria_evaluations(effect_final);
CREATE INDEX idx_crit_patient ON criteria_evaluations(patient_id);
CREATE INDEX idx_screen_elig ON screening_records(eligibility_final);
"""


def build_db(traces, out_path):
    if os.path.exists(out_path):
        os.remove(out_path)
    con = sqlite3.connect(out_path)
    con.executescript(SCHEMA)
    for tr in traces:
        pid = tr.get("patient_id", "")
        need = tr.get("patient_need") or {}
        ids = tr["_crit_ids"]
        con.execute("INSERT INTO patients VALUES (?,?,?,?,?,?)",
                    (pid, _age_band(tr), _sex(tr), need.get("need_ko"),
                     need.get("confidence"), tr.get("patient_text")))
        for t in _sorted_trials(tr):
            nct = t.get("nct_id")
            con.execute("INSERT OR IGNORE INTO trials VALUES (?,?,?)",
                        (nct, t.get("title"), t.get("phase")))
            fails, reviews, n = _counts(t)
            failed = [c for c in (t.get("criteria") or []) if c.get("effect_final") == "FAIL"]
            codes = " · ".join(ids.get((nct, c.get("text", "")), "") for c in failed)
            elig = t.get("eligibility_final")
            con.execute("INSERT INTO screening_records VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (_screening_no(pid), pid, _age_band(tr), _sex(tr), nct,
                         OUTCOME_KO.get(elig, "PENDING"), elig, codes, t.get("rank_final"),
                         fails, reviews, n, t.get("n_changed", 0), t.get("rationale_final"),
                         "미확정 (데모 범위)"))
            for c in t.get("criteria") or []:
                con.execute("INSERT INTO criteria_evaluations VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (ids.get((nct, c.get("text", "")), ""), pid, nct, c.get("text"),
                             c.get("type"), c.get("verdict_initial"), c.get("verdict_final"),
                             c.get("effect_final"), _polarity_ko(c), 1 if c.get("changed") else 0,
                             c.get("evidence"), c.get("reasoning"), c.get("action"),
                             AUTO_ACTOR, (tr.get("generated_at") or "")[:10], None, None))
        answers = _answer_map(tr)
        for q in tr.get("questions") or []:
            a = answers.get((q.get("question") or "").strip())
            if isinstance(a, (dict, list)):
                a = json.dumps(a, ensure_ascii=False)
            con.execute("INSERT INTO qa_log VALUES (?,?,?,?,?)",
                        (pid, q.get("field"), q.get("question"), q.get("why"),
                         None if a is None else str(a)))
        when = (tr.get("generated_at") or "")[:10]
        for ch in ((tr.get("reeval") or {}).get("verdict_changes")) or []:
            con.execute("INSERT INTO corrections VALUES (?,?,?,?,?,?,?,?,?)",
                        (when, pid, ids.get((ch.get("nct_id"), ch.get("criterion", "")), ""),
                         ch.get("nct_id"), ch.get("criterion"), ch.get("before"),
                         ch.get("after"), "확인 질문 응답 반영", AUTO_ACTOR))
    con.commit()
    counts = {t: con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
              for t in ("patients", "trials", "screening_records",
                        "criteria_evaluations", "qa_log", "corrections")}
    con.close()
    return counts


# --------------------------------------------------------------------------- CLI

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out-dir", default=HERE)
    ap.add_argument("--patient", help="이 환자만 (예: S001)")
    ap.add_argument("--traces", help="traces.json 경로 override")
    ap.add_argument("--no-db", action="store_true", help="SQLite는 만들지 않음")
    args = ap.parse_args()

    traces, src = load_traces(args.traces)
    digest = _sha256(src)
    if args.patient:
        traces = [t for t in traces if t.get("patient_id") == args.patient]
        if not traces:
            sys.exit(f"해당 환자를 찾을 수 없음: {args.patient}")

    stamp = datetime.date.today().strftime("%y%m%d")
    xlsx = os.path.join(args.out_dir, f"임상시험_스크리닝기록_{stamp}.xlsx")
    counts = build_workbook(traces, src, digest, xlsx)
    print(f"작성: {xlsx}")
    for k, v in counts.items():
        print(f"  {k}: {v}행")

    if not args.no_db:
        db = os.path.join(args.out_dir, "screening.db")
        dbc = build_db(traces, db)
        print(f"작성: {db}")
        for k, v in dbc.items():
            print(f"  {k}: {v}행")


if __name__ == "__main__":
    main()
