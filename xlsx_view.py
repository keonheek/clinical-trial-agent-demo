#!/usr/bin/env python3
"""Render a screening-record xlsx as one HTML page (tabs = sheets), the way the demo video
shows it -- for a browser, a slide, or a screenshot. No dependencies beyond openpyxl.

    python3 xlsx_view.py screening_record_S002.xlsx        # writes screening_record_S002.html
"""
import html
import os
import sys

from openpyxl import load_workbook

E = html.escape
GOOD = {"PASS", "MET", "적격", "충족"}
BAD = {"FAIL", "NOT_MET", "부적격", "미충족"}
WARN = {"UNKNOWN", "UNCERTAIN", "미확정", "REVIEW", "검토"}


def cell(v):
    s = "" if v is None else str(v)
    cls = "g" if s in GOOD else "b" if s in BAD else "w" if s in WARN else ""
    return f'<td class="{cls}">{E(s)}</td>'


def render(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    tabs, panes = [], []
    for i, ws in enumerate(wb.worksheets):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        width = max((max((j for j, v in enumerate(r) if v not in (None, "")), default=-1) for r in rows), default=-1) + 1
        head = "".join(f"<th>{E('' if h is None else str(h))}</th>" for h in rows[0][:width])
        body = "".join("<tr>" + "".join(cell(v) for v in r[:width]) + "</tr>" for r in rows[1:])
        tabs.append(f'<button class="tab{" on" if i == 0 else ""}" data-i="{i}">{E(ws.title)} <span>{len(rows) - 1}</span></button>')
        panes.append(f'<section id="p{i}"{"" if i == 0 else " hidden"}><h2>{E(ws.title)}</h2><div class="wrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div></section>')
    name = os.path.basename(path)
    return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8"><title>{E(name)}</title>
<style>
body{{margin:0;background:#F4F6F8;color:#11161C;font-family:"IBM Plex Sans KR","Apple SD Gothic Neo",-apple-system,sans-serif;font-size:14px}}
header{{background:#12539C;color:#fff;padding:18px 28px}} header h1{{margin:0;font-size:1.25rem;font-weight:700}} header p{{margin:4px 0 0;opacity:.8;font-size:.85rem}}
.tabs{{display:flex;gap:4px;padding:12px 28px 0;background:#fff;border-bottom:1px solid #E7EBEF}}
.tab{{font:inherit;background:#F4F6F8;border:1px solid #E7EBEF;border-bottom:none;border-radius:4px 4px 0 0;padding:8px 14px;cursor:pointer;color:#4D5967}}
.tab.on{{background:#fff;color:#12539C;font-weight:700;border-color:#D4DAE1}} .tab span{{color:#818D9B;font-size:.8rem;margin-left:4px}}
section{{padding:20px 28px}} h2{{font-size:1rem;margin:0 0 10px}}
.wrap{{overflow:auto;border:1px solid #D4DAE1;background:#fff}}
table{{border-collapse:collapse;width:100%;font-size:.84rem}}
th{{background:#1F3B63;color:#fff;text-align:left;padding:8px 10px;white-space:nowrap;position:sticky;top:0}}
td{{padding:7px 10px;border-bottom:1px solid #E7EBEF;vertical-align:top;max-width:52ch;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
td.g{{color:#1F7A4D;font-weight:600}} td.b{{color:#B3261E;font-weight:600}} td.w{{color:#9A6200;font-weight:600}}
footer{{padding:12px 28px;color:#4D5967;font-size:.8rem}}
</style></head><body>
<header><h1>{E(name)}</h1><p>export_report.py → xlsx · 시트 {len(tabs)}개 · 판정 주체는 자동 재평가(AI 파이프라인)로 기록, 확인자·확인일은 사람 검토 서명용으로 비워 둠</p></header>
<div class="tabs">{"".join(tabs)}</div>{"".join(panes)}
<footer>본 기록은 챌린지 제공 합성 환자 시나리오 기반이며 진료·등록 판단에 사용할 수 없습니다.</footer>
<script>document.querySelectorAll('.tab').forEach(b=>b.addEventListener('click',()=>{{document.querySelectorAll('.tab').forEach(x=>x.classList.remove('on'));b.classList.add('on');document.querySelectorAll('section').forEach(s=>s.hidden=s.id!=='p'+b.dataset.i);}}));</script>
</body></html>"""


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    for path in sys.argv[1:]:
        out = os.path.splitext(path)[0] + ".html"
        with open(out, "w", encoding="utf-8") as f:
            f.write(render(path))
        print("wrote", out)


if __name__ == "__main__":
    main()
